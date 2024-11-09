from typing import Generator, List, Tuple, Optional, TypeVar, Generic
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from src.pacote_log.config_log import logger
from src.dados.arquivo import Arquivo
from openpyxl.cell.cell import Cell


class ExcelDados(Arquivo[Generator[Tuple[Cell, ...], None, None], List[str],  Workbook]):
    def __init__(self, nome_arquivo=None, diretorio=None):
        super().__init__(nome_arquivo, diretorio)
        self.__planilha = self.abrir_arquivo()
        if self.__planilha is not None:
            self.__nome_aba = self.__planilha.active.title
            self.__aba = self.__planilha[self.__nome_aba]
            self.__ultima_linha = self.__aba.max_row
        else:
            self.__nome_aba = self.__aba = self.__ultima_linha = None

    def abrir_arquivo(self) -> Workbook:
        """Método para abrir a planilha"""
        planilha = load_workbook(self._caminho_arquivo)
        return planilha

    def __formatar_linhas(self, cell: Cell):
        """Método para formatar tabela"""
        cell.font = Font(bold=True, color="000000")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal="justify", vertical="center")

    def __encontrar_proxima_linha_em_branco(self, coluna: str = 'B') -> int:
        """Encontra a próxima linha em branco a partir de uma coluna específica.

        Args:
            coluna (str): A coluna onde verificar as células em branco.

        Returns:
            int: Número da próxima linha em branco.
        """
        for linha in range(2, self.__ultima_linha + 1):
            if self.__aba[f"{coluna}{linha}"].value is None:
                return linha

        return self.__ultima_linha + 1

    def ler_valores(self) -> Generator[Tuple[Cell, ...], None, None]:
        return self.__aba.iter_rows(min_row=2)

    def __formatar_tabela(self, coluna_inicial, valores, linha_para_escrever):
        for col in range(coluna_inicial, coluna_inicial + len(valores)):
            col_letter = self.__aba.cell(row=1, column=col).column_letter

            self.__aba.column_dimensions[col_letter].width = 40
        self.__aba.row_dimensions[linha_para_escrever].height = 300

    def __escrever_dados(self, valores: List[str]):
        try:

            coluna_inicial = 2
            linha_para_escrever = self.__encontrar_proxima_linha_em_branco()

            self.__formatar_tabela(coluna_inicial=coluna_inicial,
                                   valores=valores, linha_para_escrever=linha_para_escrever)

            for i, valor in enumerate(valores):

                self.__formatar_linhas(self.__planilha.active.cell(row=linha_para_escrever,
                                                                   column=coluna_inicial + i, value=valor))
            self.__ultima_linha = max(self.__ultima_linha, linha_para_escrever)
        except Exception as e:
            logger.critical(f'ERRO FATAL: {e}')
            exit()

    def gravar_dados(self, valores: List[str]):
        """_summary_

        Args:
            valores (_type_): _description_

        """

        try:
            self.__escrever_dados(valores)
            self.__planilha.save(self._caminho_arquivo)

        except OSError as e:
            logger.error(f'Erro de sistema ao salvar o arquivo: {e}')
            exit()
        except Exception as e:
            logger.error(f'Erro ao salvar o arquivo: {e}')
            exit()

    def __del__(self):
        self.__planilha.close()
